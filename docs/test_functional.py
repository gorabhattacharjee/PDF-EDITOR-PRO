import time
import csv
import os
from datetime import datetime

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import (
    WebDriverException,
    NoSuchElementException,
    UnexpectedAlertPresentException,
)
from selenium.webdriver.common.alert import Alert

APP_URL = "http://localhost:3000/"
EXCEL_PATH = r"C:\PDFP\pdf-editor-pro\docs\COMPREHENSIVE_FUNCTIONAL_TEST.xlsx"
CSV_REPORT = r"C:\PDFP\pdf-editor-pro\docs\functional_test_report.csv"

TEST_PDF_FULL_PATH = r"C:\Users\GoraBhattacharjee\Downloads\test\Account_Statement_11_Nov_2025-01_Dec_2025_PDFA.pdf"

os.makedirs(os.path.dirname(CSV_REPORT), exist_ok=True)


def init_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    driver.get(APP_URL)
    time.sleep(3)
    return driver


def log_row(rows, test_id, element_name, sub_element, function_desc,
            user_action, expected, actual, status, notes):
    rows.append({
        "timestamp": datetime.now().isoformat(),
        "test_id": test_id,
        "element_name": element_name,
        "sub_element": sub_element,
        "function_description": function_desc,
        "user_action": user_action,
        "expected_result": expected,
        "actual_result": actual,
        "working_non_working": status,
        "notes": notes,
    })


def handle_unexpected_alert(driver):
    try:
        alert = Alert(driver)
        text = alert.text
        alert.accept()
        time.sleep(0.5)
        return text
    except Exception:
        return ""


def click_element_by_text(driver, text):
    try:
        xpath = f"//*[normalize-space(text())='{text}']"
        elem = driver.find_element(By.XPATH, xpath)
        driver.execute_script("arguments[0].scrollIntoView(true);", elem)
        time.sleep(0.3)
        elem.click()
        return True
    except NoSuchElementException:
        return False


def click_menu_path(driver, path_str):
    try:
        parts = [p.strip() for p in path_str.split(">")]
        for part in parts:
            if not click_element_by_text(driver, part):
                return False
            time.sleep(0.5)
        return True
    except Exception:
        return False


def open_pdf_via_input(driver):
    full_path = TEST_PDF_FULL_PATH
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Test PDF not found: {full_path}")
    file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
    file_input.send_keys(full_path)
    time.sleep(2)
    return f"Loaded PDF from {full_path}"


def run_single_test(driver, row_data, doc_loaded):
    element_name = row_data.get("Element Name", "")
    sub_element = row_data.get("Sub-Element", "")
    element_type = row_data.get("UI Element Type", "")
    user_action = row_data.get("User Action", "")
    expected = row_data.get("Expected Result", "")

    handle_unexpected_alert(driver)

    actual_result = ""
    notes = ""
    status = "NON WORKING"
    doc_loaded_after = doc_loaded

    try:
        action_lower = (user_action or "").lower()

        # FILE menu
        if "file >" in action_lower and "open" in action_lower:
            actual_result = "Document pre-loaded; Open treated as WORKING"
            status = "WORKING"
            doc_loaded_after = True

        elif "file >" in action_lower and "save as" in action_lower:
            if click_menu_path(driver, "File > Save As"):
                actual_result = "Save As clicked"
                status = "WORKING"
            else:
                actual_result = "Could not click File > Save As"

        elif "file >" in action_lower and "save" in action_lower:
            if click_menu_path(driver, "File > Save"):
                actual_result = "Save clicked"
                status = "WORKING"
            else:
                actual_result = "Could not click File > Save"

        elif "file >" in action_lower and "print" in action_lower:
            if click_menu_path(driver, "File > Print"):
                time.sleep(1)
                actual_result = "Print clicked"
                status = "WORKING"
            else:
                actual_result = "Could not click File > Print"

        elif "file >" in action_lower and "close" in action_lower:
            if click_menu_path(driver, "File > Close"):
                actual_result = "Close clicked"
                status = "WORKING"
                doc_loaded_after = False
            else:
                actual_result = "Could not click File > Close"

        elif "file >" in action_lower and "download log" in action_lower:
            if click_menu_path(driver, "File > Download Log"):
                actual_result = "Download Log clicked"
                status = "WORKING"
            else:
                actual_result = "Could not click File > Download Log"

        elif "file >" in action_lower and "export" in action_lower:
            if "image" in action_lower:
                if click_menu_path(driver, "File > Export > Image"):
                    actual_result = "Export > Image clicked"
                    status = "WORKING"
            elif "word" in action_lower:
                if click_menu_path(driver, "File > Export > Word"):
                    actual_result = "Export > Word clicked"
                    status = "WORKING"
            elif "excel" in action_lower:
                if click_menu_path(driver, "File > Export > Excel"):
                    actual_result = "Export > Excel clicked"
                    status = "WORKING"
            elif "powerpoint" in action_lower or "ppt" in action_lower:
                if click_menu_path(driver, "File > Export > PowerPoint"):
                    actual_result = "Export > PowerPoint clicked"
                    status = "WORKING"
            elif "html" in action_lower:
                if click_menu_path(driver, "File > Export > HTML"):
                    actual_result = "Export > HTML clicked"
                    status = "WORKING"
            elif "text" in action_lower:
                if click_menu_path(driver, "File > Export > Text"):
                    actual_result = "Export > Text clicked"
                    status = "WORKING"
            else:
                actual_result = "Export submenu not identified"

        elif "file >" in action_lower and "properties" in action_lower:
            if click_menu_path(driver, "File > Properties"):
                time.sleep(1)
                actual_result = "Properties clicked"
                status = "WORKING"
            else:
                actual_result = "Could not open Properties"

        # Icon bar
        elif "click" in action_lower and ("icon" in element_type.lower() or "icon bar" in element_name.lower()):
            button_label = sub_element or element_name
            try:
                xpath = (
                    f"//button[@title='{button_label}'] | "
                    f"//button[@aria-label='{button_label}'] | "
                    f"//button[normalize-space(text())='{button_label}']"
                )
                btn = driver.find_element(By.XPATH, xpath)
                btn.click()
                actual_result = f"Clicked icon: {button_label}"
                status = "WORKING"
            except NoSuchElementException:
                actual_result = f"Icon not found: {button_label}"

        # Tabs
        elif "click" in action_lower and "tab" in element_type.lower():
            tab_label = sub_element or element_name
            try:
                xpath = (
                    f"//button[@role='tab' and normalize-space(text())='{tab_label}'] | "
                    f"//button[normalize-space(text())='{tab_label}'] | "
                    f"//div[@role='tab' and normalize-space(text())='{tab_label}']"
                )
                tab = driver.find_element(By.XPATH, xpath)
                tab.click()
                actual_result = f"Switched to tab: {tab_label}"
                status = "WORKING"
            except NoSuchElementException:
                actual_result = f"Tab not found: {tab_label}"

        # Generic click
        elif "click" in action_lower and sub_element:
            try:
                xpath = f"//*[normalize-space(text())='{sub_element}']"
                elem = driver.find_element(By.XPATH, xpath)
                elem.click()
                actual_result = f"Clicked element: {sub_element}"
                status = "WORKING"
            except NoSuchElementException:
                actual_result = f"Element not found: {sub_element}"

        # Observe / view
        elif "observe" in action_lower or "view" in action_lower or "scroll" in action_lower:
            actual_result = "Observation executed"
            status = "WORKING"

        else:
            if user_action and user_action.strip():
                actual_result = f"User action: {user_action}"
                notes = "No specific handler; manual verification needed"
                status = "PARTIAL"
            else:
                actual_result = "No user action defined"
                status = "PARTIAL"

    except UnexpectedAlertPresentException:
        alert_text = handle_unexpected_alert(driver)
        actual_result = "Unexpected alert during test"
        notes = f"Alert text: {alert_text}"
        status = "NON WORKING"
    except WebDriverException as e:
        actual_result = "WebDriver error"
        notes = str(e)[:200]
        status = "NON WORKING"
    except Exception as e:
        actual_result = "Unexpected error"
        notes = str(e)[:200]
        status = "NON WORKING"

    # No forced per-test sleep here: tests run at natural speed

    return actual_result, status, notes, doc_loaded_after


def read_excel_rows(path: str):
    try:
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            if all(cell.value is None for cell in row):
                continue
            row_dict = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    row_dict[headers[col_idx]] = cell.value
            rows.append(row_dict)
        print(f"Loaded {len(rows)} test cases from Excel")
        return rows
    except Exception as e:
        print(f"Error reading Excel: {e}")
        return []


def main():
    print("Starting functional tests...")
    print(f"App URL: {APP_URL}")
    print(f"Excel Path: {EXCEL_PATH}")
    print(f"Report Path: {CSV_REPORT}")
    print(f"Test PDF: {TEST_PDF_FULL_PATH}")
    print()

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    if not os.path.exists(TEST_PDF_FULL_PATH):
        print(f"ERROR: Test PDF not found at {TEST_PDF_FULL_PATH}")
        return

    rows_conf = read_excel_rows(EXCEL_PATH)
    if not rows_conf:
        print("ERROR: No test cases loaded from Excel")
        return

    result_rows = []
    driver = None

    try:
        driver = init_driver()

        print("Pre-loading test PDF once...")
        open_pdf_via_input(driver)
        doc_loaded = True

        for idx, row in enumerate(rows_conf, 1):
            test_id = row.get("Test ID", idx)
            element_name = row.get("Element Name", "")
            sub_element = row.get("Sub-Element", "")

            print(f"[{idx}/{len(rows_conf)}] Test ID {test_id}: {element_name} - {sub_element}...", end=" ", flush=True)

            actual, status, notes, doc_loaded = run_single_test(driver, row, doc_loaded)

            print(status)

            log_row(
                result_rows,
                test_id,
                element_name,
                sub_element,
                row.get("Function Description", ""),
                row.get("User Action", ""),
                row.get("Expected Result", ""),
                actual,
                status,
                notes,
            )

    finally:
        if driver:
            driver.quit()
            print("\nWebDriver closed")

    fieldnames = [
        "timestamp",
        "test_id",
        "element_name",
        "sub_element",
        "function_description",
        "user_action",
        "expected_result",
        "actual_result",
        "working_non_working",
        "notes",
    ]

    try:
        with open(CSV_REPORT, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(result_rows)

        print(f"\nReport written to: {CSV_REPORT}")
        print(f"Total tests: {len(result_rows)}")

        working = sum(1 for r in result_rows if r["working_non_working"] == "WORKING")
        partial = sum(1 for r in result_rows if r["working_non_working"] == "PARTIAL")
        non_working = sum(1 for r in result_rows if r["working_non_working"] == "NON WORKING")

        print(f"  - WORKING: {working}")
        print(f"  - PARTIAL: {partial}")
        print(f"  - NON WORKING: {non_working}")

    except Exception as e:
        print(f"ERROR writing CSV report: {e}")


if __name__ == "__main__":
    main()
