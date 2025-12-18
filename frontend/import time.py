import time
import csv
import os
from datetime import datetime

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException, NoSuchElementException, TimeoutException

APP_URL = "http://localhost:3000/"
EXCEL_PATH = r"C:\\pdf-editor-pro\\docs\\COMPREHENSIVE_FUNCTIONAL_TEST.xlsx"
TEST_FILES_PATH = r"C:\\pdf-editor-pro\\test"
CSV_REPORT = r"C:\\pdf-editor-pro\\docs\\functional_test_report.csv"
TEST_PDF_NAME = "sample.pdf"

os.makedirs(os.path.dirname(CSV_REPORT), exist_ok=True)

def init_driver():
    """Initialize Selenium WebDriver with Chrome options."""
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    # options.add_argument("--headless=new")
    try:
        driver = webdriver.Chrome(options=options)
        driver.get(APP_URL)
        time.sleep(3)
        return driver
    except WebDriverException as e:
        print(f"Error initializing WebDriver: {e}")
        print("Please ensure ChromeDriver is installed and in your system's PATH.")
        return None

def log_row(rows, test_id, element_name, sub_element, function_desc,
            user_action, expected, actual, status, notes):
    """Log a test result row."""
    rows.append({
        "timestamp": datetime.now().isoformat(),
        "test_id": test_id,
        "element_name": element_name,
        "sub_element": sub_element,
        "function_description": function_desc,
        "user_action": user_action,
        "expected_result": expected,
        "actual_result": actual,
        "working_non_working": status,
        "notes": notes,
    })

def click_element_by_text(driver, text):
    """Click an element by its visible text."""
    if not text or not isinstance(text, str):
        return False
    try:
        xpath = f"//*[normalize-space(text())='{text}']"
        elem = driver.find_element(By.XPATH, xpath)
        driver.execute_script("arguments[0].scrollIntoView(true);", elem)
        time.sleep(0.3)
        elem.click()
        return True
    except NoSuchElementException:
        print(f"Element with text '{text}' not found.")
        return False
    except Exception as e:
        print(f"Error clicking element with text '{text}': {e}")
        return False

def click_menu_path(driver, path_str):
    """Click a menu path like 'File > Open' by visible text."""
    if not path_str or not isinstance(path_str, str):
        return False
    try:
        parts = [p.strip() for p in path_str.split(">")]
        for part in parts:
            if not click_element_by_text(driver, part):
                return False
            time.sleep(0.5)
        return True
    except Exception as e:
        print(f"Error clicking menu path '{path_str}': {e}")
        return False

def run_single_test(driver, row_data):
    """
    Execute a single test case based on Excel row data.
    """
    element_name = row_data.get("Element Name")
    sub_element = row_data.get("Sub-Element")
    
    action_path = str(element_name) if not sub_element or str(sub_element).isspace() else f"{element_name} > {sub_element}"
    
    driver.refresh()
    time.sleep(1)

    try:
        if click_menu_path(driver, action_path):
            return "Working", f"Successfully clicked on '{action_path}'."
        else:
            return "Non-working", f"Failed to click on '{action_path}'."
    except Exception as e:
        return "Non-working", f"An error occurred during test: {e}"

def main():
    """
    Main function to run the test suite.
    """
    driver = init_driver()
    if not driver:
        return

    rows = []
    
    try:
        try:
            workbook = load_workbook(EXCEL_PATH)
            sheet = workbook.active
        except FileNotFoundError:
            print(f"Error: Test data file not found at {EXCEL_PATH}")
            return

        header = [cell.value for cell in sheet[1]]
        
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            row_data = dict(zip(header, row))
            print(f"Testing item {i-1}: {row_data.get('Element Name')} > {row_data.get('Sub-Element')}")
            
            status, notes = run_single_test(driver, row_data)
            
            log_row(
                rows,
                row_data.get("Test ID"),
                row_data.get("Element Name"),
                row_data.get("Sub-Element"),
                row_data.get("Function to be implemented"),
                row_data.get("User Action"),
                row_data.get("Expected Result"),
                "",
                status,
                notes
            )
            
    finally:
        if driver:
            driver.quit()
        
        if rows:
            print(f"\\nWriting test results to {CSV_REPORT}")
            try:
                with open(CSV_REPORT, "w", newline="", encoding="utf-8") as f:
                    fieldnames = list(rows[0].keys())
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
                print("Test report generated successfully.")
            except Exception as e:
                print(f"Error writing report: {e}")
        else:
            print("No test results to write.")

if __name__ == "__main__":
    main()
